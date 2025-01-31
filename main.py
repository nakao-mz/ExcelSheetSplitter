import tkinter.messagebox
import openpyxl
from openpyxl import Workbook
from tkinter import Tk, Label, Button, Radiobutton, StringVar
from tkinter.filedialog import askopenfilename
import os


def get_excel_file_path():
    """
    Excelファイルを選択するためのダイアログを表示し、選択されたファイルのパスを返す関数。
    """
    Tk().withdraw()  # ルートウィンドウを表示しないように設定
    file_path = askopenfilename()  # ファイル選択ダイアログを表示し、選択されたファイルのパスを取得
    if not file_path:  # ファイルが選択されなかった場合
        print("ファイルが選択されませんでした。")
    return file_path  # 選択されたファイルのパスを返す


def create_new_excel_with_sheets(mode):
    """
    Excelファイルを読み込み、特定のシートの値を基に新しいExcelファイルを作成する関数。
    """
    original_excel_file_path = get_excel_file_path()  # 元のExcelファイルのパスを取得
    if not original_excel_file_path:  # ファイルが選択されなかった場合は処理を終了
        return

    original_workbook = openpyxl.load_workbook(original_excel_file_path, data_only=True)  # 元のExcelファイルを読み込む（数式ではなく値を取得）
    original_sheet = original_workbook["試験項目"]  # "試験項目"シートを取得

    if mode == "single":
        new_workbook = Workbook()  # 新しいExcelブックを作成
        new_workbook.remove(new_workbook.active)  # デフォルトで作成されるシートを削除

        for row_num in range(8, original_sheet.max_row + 1):  # 8行目から最終行までループ
            cell_value = original_sheet.cell(row=row_num, column=1).value  # 1列目の値を取得
            if cell_value is None:  # 値が空の場合はループを抜ける
                break
            new_sheet_title = f"{cell_value}"  # 新しいシートのタイトルを設定
            new_workbook.create_sheet(title=new_sheet_title)  # 新しいシートを作成

        original_file_name = os.path.basename(original_excel_file_path)  # 元のファイル名を取得
        original_dir_path = os.path.dirname(original_excel_file_path)  # 元のファイルのディレクトリパスを取得
        new_file_name = f"エビ_{original_file_name}"  # 新しいファイル名を作成

        new_file_path = os.path.join(original_dir_path, new_file_name)  # 新しいファイルのパスを作成
        new_workbook.save(new_file_path)  # 新しいExcelファイルを保存
        tkinter.messagebox.showinfo("完了", f"新しいExcelファイルを作成しました。\n{new_file_path}")  # 完了メッセージを表示

    elif mode == "multiple":
        test_item_dict = {}
        previous_test_item = None

        for row_num in range(8, original_sheet.max_row + 1):
            test_number = original_sheet.cell(row=row_num, column=1).value
            test_item = original_sheet.cell(row=row_num, column=5).value
            end_value = original_sheet.cell(row=row_num, column=8).value

            if end_value is None:
                break

            if test_item is None:
                test_item = previous_test_item
            else:
                previous_test_item = test_item

            if test_item not in test_item_dict:
                test_item_dict[test_item] = []

            test_item_dict[test_item].append(test_number)

        original_dir_path = os.path.dirname(original_excel_file_path)

        for test_item, test_numbers in test_item_dict.items():
            start_number = test_numbers[0]
            end_number = test_numbers[-1]
            new_file_name = f"No.{start_number}~{end_number}_エビデンス_{test_item}.xlsx"
            new_file_path = os.path.join(original_dir_path, new_file_name)

            new_workbook = Workbook()
            new_workbook.remove(new_workbook.active)

            for test_number in test_numbers:
                new_sheet_title = f"{test_number}"
                new_workbook.create_sheet(title=new_sheet_title)

            new_workbook.save(new_file_path)

        tkinter.messagebox.showinfo("完了", "新しいExcelファイルを作成しました。")  # 完了メッセージを表示

    root = Tk()
    root.protocol("WM_DELETE_WINDOW", root.quit)
    root.quit()


def select_mode():
    """
    ファイル作成モードを選択するためのGUIを表示する関数。
    """
    def on_select():
        mode = mode_var.get()
        if mode in ["single", "multiple"]:
            root.destroy()
            create_new_excel_with_sheets(mode)
        else:
            tkinter.messagebox.showerror("エラー", "無効なモードが選択されました。")

    root = Tk()
    root.title("ファイル作成モード選択")

    Label(root, text="ファイル作成モードを選択してください:").pack()

    mode_var = StringVar(value="single")
    Radiobutton(root, text="Single", variable=mode_var, value="single").pack()
    Radiobutton(root, text="Multiple", variable=mode_var, value="multiple").pack()

    Button(root, text="OK", command=on_select).pack()

    root.protocol("WM_DELETE_WINDOW", root.quit)
    root.mainloop()


if __name__ == "__main__":  # このスクリプトが直接実行された場合のみ以下の処理を実行
    select_mode()  # ファイル作成モード選択のGUIを表示
